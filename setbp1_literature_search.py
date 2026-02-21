#!/usr/bin/env python3
"""
SETBP1 Literature Search Script
Searches PubMed, bioRxiv, and medRxiv for SETBP1 and Schinzel-Giedion Syndrome papers
Generates Excel and PDF reports with comprehensive citations

Usage:
    python setbp1_literature_search.py --start 2026-02-01 --end 2026-02-08
    python setbp1_literature_search.py  # Uses last 7 days by default
"""

import argparse
import json
import time
from datetime import datetime, timedelta
from typing import List, Dict, Any
import requests
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, PageBreak, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_JUSTIFY


class SETBPLiteratureSearch:
    """Comprehensive SETBP1 literature search across multiple databases"""
    
    def __init__(self, start_date: str, end_date: str, output_dir: str = "/mnt/user-data/outputs"):
        self.start_date = start_date
        self.end_date = end_date
        self.output_dir = output_dir
        self.all_papers = []
        self.pubmed_base = "https://eutils.ncbi.nlm.nih.gov/entrez/eutils/"
        
    def search_pubmed(self) -> List[str]:
        """Search PubMed for SETBP1 papers"""
        print(f"Searching PubMed for papers from {self.start_date} to {self.end_date}...")
        
        # Build search query
        query = '("SETBP1" OR "Schinzel-Giedion Syndrome" OR "Schinzel Giedion" OR "SET binding protein 1")'
        
        # Search parameters
        params = {
            'db': 'pubmed',
            'term': query,
            'retmax': 1000,
            'retmode': 'json',
            'datetype': 'pdat',
            'mindate': self.start_date.replace('-', '/'),
            'maxdate': self.end_date.replace('-', '/')
        }
        
        try:
            response = requests.get(f"{self.pubmed_base}esearch.fcgi", params=params)
            response.raise_for_status()
            data = response.json()
            pmids = data.get('esearchresult', {}).get('idlist', [])
            print(f"  Found {len(pmids)} papers in PubMed")
            return pmids
        except Exception as e:
            print(f"  Error searching PubMed: {e}")
            return []
    
    def get_pubmed_metadata(self, pmids: List[str], batch_size: int = 100) -> List[Dict]:
        """Retrieve metadata for PubMed articles in batches with retry logic"""
        print(f"Retrieving metadata for {len(pmids)} PubMed articles...")
        all_metadata = []
        
        for i in range(0, len(pmids), batch_size):
            batch = pmids[i:i + batch_size]
            batch_num = (i // batch_size) + 1
            total_batches = (len(pmids) + batch_size - 1) // batch_size
            
            print(f"  Processing batch {batch_num}/{total_batches} ({len(batch)} PMIDs)...")
            
            max_retries = 3
            for attempt in range(max_retries):
                try:
                    summary_params = {
                        'db': 'pubmed',
                        'id': ','.join(batch),
                        'retmode': 'json'
                    }
                    
                    summary_response = requests.get(f"{self.pubmed_base}esummary.fcgi", params=summary_params)
                    summary_response.raise_for_status()
                    summary_data = summary_response.json()
                    
                    for pmid in batch:
                        if pmid in summary_data.get('result', {}):
                            article = summary_data['result'][pmid]
                            
                            # Get full author list (not truncated)
                            all_authors = [a.get('name', '') for a in article.get('authors', [])]
                            authors_string = ', '.join(all_authors) if all_authors else ''
                            
                            metadata = {
                                'pmid': pmid,
                                'title': article.get('title', ''),
                                'authors': authors_string,
                                'journal': article.get('source', ''),
                                'year': article.get('pubdate', '').split()[0] if article.get('pubdate') else '',
                                'doi': next((id['value'] for id in article.get('articleids', []) 
                                           if id.get('idtype') == 'doi'), ''),
                                'source': 'PubMed'
                            }
                            all_metadata.append(metadata)
                    
                    print(f"    Batch {batch_num} completed successfully")
                    break  # Success, exit retry loop
                    
                except Exception as e:
                    if attempt < max_retries - 1:
                        wait_time = (attempt + 1) * 2
                        print(f"    Attempt {attempt + 1} failed: {e}")
                        print(f"    Retrying in {wait_time} seconds...")
                        time.sleep(wait_time)
                    else:
                        print(f"    Batch {batch_num} failed after {max_retries} attempts: {e}")
            
            # Rate limiting
            time.sleep(0.5)
        
        print(f"  Successfully retrieved metadata for {len(all_metadata)} articles")
        return all_metadata
    
    # Keywords used to filter bioRxiv/medRxiv results for PKD relevance
    SETBP1_KEYWORDS = [
        'setbp1', 'schinzel-giedion', 'schinzel giedion', 'sgd',
        'set binding protein 1',
    ]

    def _is_setbp1_relevant(self, title: str, abstract: str = '') -> bool:
        """Check if a paper is SETBP1-relevant based on title and abstract"""
        text = (title + ' ' + abstract).lower()
        return any(kw in text for kw in self.SETBP1_KEYWORDS)

    def _search_preprint_server(self, server: str) -> List[Dict]:
        """Search bioRxiv or medRxiv for PKD preprints via the public API.

        The API (api.biorxiv.org) returns all preprints in a date range.
        We fetch pages of 100 and filter locally for PKD-relevant papers.
        """
        base_url = f"https://api.biorxiv.org/details/{server}"
        papers = []
        cursor = 0
        page_size = 100

        while True:
            url = f"{base_url}/{self.start_date}/{self.end_date}/{cursor}"
            try:
                resp = requests.get(url, timeout=30)
                resp.raise_for_status()
                data = resp.json()
            except Exception as e:
                print(f"    API error at cursor {cursor}: {e}")
                break

            collection = data.get('collection', [])
            if not collection:
                break

            for item in collection:
                title = item.get('title', '')
                abstract = item.get('abstract', '')
                if self._is_setbp1_relevant(title, abstract):
                    # Build author string: "Last1 FI, Last2 FI, ..."
                    authors_raw = item.get('authors', '')

                    doi = item.get('doi', '')
                    papers.append({
                        'pmid': '',  # preprints don't have PMIDs
                        'title': title,
                        'authors': authors_raw,
                        'journal': f"{server} (preprint)",
                        'year': item.get('date', '')[:4],
                        'doi': doi,
                        'source': server,
                    })

            # Check if there are more pages
            total_msg = data.get('messages', [{}])[0]
            total_count = int(total_msg.get('total', 0)) if total_msg else 0
            cursor += page_size
            if cursor >= total_count:
                break

            time.sleep(0.3)  # rate limiting

        return papers

    def search_biorxiv(self) -> List[Dict]:
        """Search bioRxiv for SETBP1 preprints"""
        print(f"Searching bioRxiv for preprints from {self.start_date} to {self.end_date}...")
        papers = self._search_preprint_server('biorxiv')
        print(f"  Found {len(papers)} SETBP1-relevant preprints in bioRxiv")
        return papers

    def search_medrxiv(self) -> List[Dict]:
        """Search medRxiv for SETBP1 preprints"""
        print(f"Searching medRxiv for preprints from {self.start_date} to {self.end_date}...")
        papers = self._search_preprint_server('medrxiv')
        print(f"  Found {len(papers)} SETBP1-relevant preprints in medRxiv")
        return papers
    
    def categorize_papers(self, papers: List[Dict]) -> Dict[str, List[Dict]]:
        """Categorize papers by research area"""
        categories = {
            'mechanism': [],
            'therapeutics': [],
            'models': [],
            'dataset': [],
            'other': []
        }

        # Simple keyword-based categorization
        for paper in papers:
            title_lower = paper.get('title', '').lower()

            if any(word in title_lower for word in ['mechanism', 'pathway', 'signaling', 'function', 'structure']):
                categories['mechanism'].append(paper)
            elif any(word in title_lower for word in ['drug', 'therapeutic', 'treatment', 'inhibitor', 'trial']):
                categories['therapeutics'].append(paper)
            elif any(word in title_lower for word in ['mouse', 'mice', 'rat', 'model', 'crispr', 'cell', 'in vitro']):
                categories['models'].append(paper)
            elif any(word in title_lower for word in ['cohort', 'registry', 'dataset', 'population']):
                categories['dataset'].append(paper)
            else:
                categories['other'].append(paper)

        return categories
    
    def create_summary(self, title: str, max_words: int = 8) -> str:
        """Create concise summary from title"""
        # Remove common words and create short summary
        words = title.split()
        
        # Take first max_words
        summary_words = words[:max_words] if len(words) <= max_words else words[:max_words-1]
        
        # Clean up
        summary = ' '.join(summary_words)
        if len(words) > max_words:
            summary = summary.rstrip('.,;:') + '...'
        
        return summary
    
    def create_key_findings(self, title: str, abstract: str = "", max_words: int = 20) -> str:
        """Create key findings summary (20 words or less)"""
        # For now, use a simplified version from the title
        # In production, you'd parse the abstract
        words = title.replace(':', ' ').split()
        
        # Remove very common words
        skip = {'the', 'a', 'an', 'in', 'on', 'at', 'to', 'for', 'of', 'and', 'with'}
        filtered = [w for w in words if w.lower() not in skip or len([x for x in words if x.lower() not in skip]) < 10]
        
        # Take up to max_words
        key_words = filtered[:max_words]
        findings = ' '.join(key_words)
        
        return findings
    
    def extract_last_author_name(self, authors_string: str) -> str:
        """Extract last author's last name from author string"""
        if not authors_string:
            return ""
        
        # Split by comma to get individual authors
        authors = [a.strip() for a in authors_string.split(',')]
        if not authors:
            return ""
        
        # Get last author
        last_author = authors[-1]
        
        # Extract last name (first word before any space)
        parts = last_author.split()
        return parts[0] if parts else ""
    
    def create_excel_report(self, papers: List[Dict], filename: str):
        """Create Excel report with specific column format
        
        Format:
        A-C: Blank columns
        D: Summary (less than 8 words)
        E: Last author's last name
        F: Journal
        G: Key findings (20 words or less)
        H-I: Blank columns
        J: Web link
        """
        print(f"Creating Excel report: {filename}...")
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "SETBP1 Literature"
        
        # Set column widths
        ws.column_dimensions['A'].width = 5   # Blank
        ws.column_dimensions['B'].width = 5   # Blank
        ws.column_dimensions['C'].width = 5   # Blank
        ws.column_dimensions['D'].width = 35  # Summary
        ws.column_dimensions['E'].width = 15  # Last Author
        ws.column_dimensions['F'].width = 25  # Journal
        ws.column_dimensions['G'].width = 60  # Key Findings
        ws.column_dimensions['H'].width = 5   # Blank
        ws.column_dimensions['I'].width = 5   # Blank
        ws.column_dimensions['J'].width = 45  # Link
        
        # Header formatting
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Set headers
        headers = ["", "", "", "Summary", "Last Author", "Journal", "Key Findings", "", "", "Link"]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            if header:  # Only style non-blank headers
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
        
        # Add data
        for idx, paper in enumerate(papers, start=2):
            # Columns A, B, C - blank
            ws.cell(row=idx, column=1, value="")
            ws.cell(row=idx, column=2, value="")
            ws.cell(row=idx, column=3, value="")
            
            # Column D - Summary (less than 8 words)
            summary = self.create_summary(paper.get('title', ''), max_words=7)
            ws.cell(row=idx, column=4, value=summary)
            
            # Column E - Last author's last name
            last_author = self.extract_last_author_name(paper.get('authors', ''))
            ws.cell(row=idx, column=5, value=last_author)
            
            # Column F - Journal
            ws.cell(row=idx, column=6, value=paper.get('journal', ''))
            
            # Column G - Key findings (20 words or less)
            key_findings = self.create_key_findings(paper.get('title', ''), max_words=20)
            ws.cell(row=idx, column=7, value=key_findings)
            
            # Columns H, I - blank
            ws.cell(row=idx, column=8, value="")
            ws.cell(row=idx, column=9, value="")
            
            # Column J - Web link
            doi = paper.get('doi', '')
            pmid = paper.get('pmid', '')
            
            if doi:
                link = f"https://doi.org/{doi}"
            elif pmid:
                link = f"https://pubmed.ncbi.nlm.nih.gov/{pmid}/"
            else:
                link = ""
            
            if link:
                link_cell = ws.cell(row=idx, column=10)
                link_cell.value = link
                link_cell.hyperlink = link
                link_cell.font = Font(color="0563C1", underline="single")
        
        # Save Excel
        wb.save(filename)
        print(f"  Excel report saved: {filename}")
        print(f"  Format: A-C blank | D summary | E last author | F journal | G key findings | H-I blank | J link")
    
    def _paper_link(self, paper: Dict) -> str:
        """Return the best available link for a paper (DOI or PubMed)"""
        doi = paper.get('doi', '')
        pmid = paper.get('pmid', '')
        if doi:
            return f'https://doi.org/{doi}'
        elif pmid:
            return f'https://pubmed.ncbi.nlm.nih.gov/{pmid}/'
        return ''

    def _paper_link_label(self, paper: Dict) -> str:
        """Return a formatted link label for PDF citations"""
        doi = paper.get('doi', '')
        pmid = paper.get('pmid', '')
        link = self._paper_link(paper)
        if not link:
            return ''
        if pmid:
            return (f'<link href="{link}" color="blue">'
                    f'PMID: {pmid}</link>')
        elif doi:
            return (f'<link href="{link}" color="blue">'
                    f'DOI: {doi}</link>')
        return ''

    def create_pdf_report(self, papers: List[Dict], categories: Dict, filename: str,
                          pubmed_count: int = 0, biorxiv_count: int = 0, medrxiv_count: int = 0):
        """Create comprehensive PDF report"""
        print(f"Creating PDF report: {filename}...")
        
        doc = SimpleDocTemplate(filename, pagesize=letter,
                              rightMargin=72, leftMargin=72,
                              topMargin=72, bottomMargin=18)
        
        story = []
        styles = getSampleStyleSheet()
        
        # Custom styles
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            textColor=colors.HexColor('#1f4788'),
            spaceAfter=30,
            alignment=TA_CENTER
        )
        
        heading1_style = ParagraphStyle(
            'CustomHeading1',
            parent=styles['Heading1'],
            fontSize=14,
            textColor=colors.HexColor('#1f4788'),
            spaceAfter=12,
            spaceBefore=12,
            borderWidth=1,
            borderColor=colors.HexColor('#1f4788'),
            borderPadding=5
        )
        
        heading2_style = ParagraphStyle(
            'CustomHeading2',
            parent=styles['Heading2'],
            fontSize=12,
            textColor=colors.HexColor('#2E5C8A'),
            spaceAfter=10,
            spaceBefore=10
        )
        
        body_style = ParagraphStyle(
            'CustomBody',
            parent=styles['Normal'],
            fontSize=10,
            leading=14,
            alignment=TA_JUSTIFY
        )
        
        citation_style = ParagraphStyle(
            'Citation',
            parent=styles['Normal'],
            fontSize=9,
            leftIndent=20,
            textColor=colors.HexColor('#444444'),
            leading=12,
            spaceAfter=8
        )
        
        # Title
        story.append(Paragraph("SETBP1 LITERATURE REVIEW", title_style))
        story.append(Paragraph(f"Search Period: {self.start_date} to {self.end_date}", styles['Normal']))
        story.append(Spacer(1, 20))
        
        # Search Summary
        story.append(Paragraph("SEARCH SUMMARY", heading1_style))
        summary_data = [
            ["PubMed Articles:", str(pubmed_count)],
            ["bioRxiv Preprints:", str(biorxiv_count)],
            ["medRxiv Preprints:", str(medrxiv_count)],
            ["Total Papers:", str(len(papers))],
            ["Date Range:", f"{self.start_date} to {self.end_date}"],
        ]
        summary_table = Table(summary_data, colWidths=[2.5*inch, 3*inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#E8EFF7')),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey)
        ]))
        story.append(summary_table)
        story.append(Spacer(1, 20))
        
        # Notable Findings by Category
        story.append(Paragraph("NOTABLE FINDINGS", heading1_style))
        story.append(Spacer(1, 10))

        # Mechanism
        if categories['mechanism']:
            story.append(Paragraph("MECHANISM:", heading2_style))
            for paper in categories['mechanism'][:3]:
                story.append(Paragraph(f"• {paper['title']}", body_style))
                citation = (f'<b>Citation:</b> {paper["authors"]}. '
                          f'{self._paper_link_label(paper)}')
                story.append(Paragraph(citation, citation_style))
            story.append(Spacer(1, 15))

        # Therapeutics
        if categories['therapeutics']:
            story.append(Paragraph("THERAPEUTICS:", heading2_style))
            for paper in categories['therapeutics'][:4]:
                story.append(Paragraph(f"• {paper['title']}", body_style))
                citation = (f'<b>Citation:</b> {paper["authors"]}. '
                          f'{self._paper_link_label(paper)}')
                story.append(Paragraph(citation, citation_style))
            story.append(Spacer(1, 15))

        # Models
        if categories['models']:
            story.append(Paragraph("MODELS:", heading2_style))
            for paper in categories['models'][:4]:
                story.append(Paragraph(f"• {paper['title']}", body_style))
                citation = (f'<b>Citation:</b> {paper["authors"]}. '
                          f'{self._paper_link_label(paper)}')
                story.append(Paragraph(citation, citation_style))
            story.append(Spacer(1, 15))

        # New Data Sets
        if categories['dataset']:
            story.append(Paragraph("NEW DATA SETS:", heading2_style))
            for paper in categories['dataset'][:3]:
                story.append(Paragraph(f"• {paper['title']}", body_style))
                citation = (f'<b>Citation:</b> {paper["authors"]}. '
                          f'{self._paper_link_label(paper)}')
                story.append(Paragraph(citation, citation_style))
        
        story.append(PageBreak())
        
        # Complete Paper List
        story.append(Paragraph(f"COMPLETE PAPER LIST ({len(papers)} PAPERS)", heading1_style))
        story.append(Spacer(1, 15))
        
        for i, paper in enumerate(papers, 1):
            paper_text = f'<b>{i}. {paper["title"]}</b>'
            story.append(Paragraph(paper_text, body_style))
            
            citation_text = f'{paper["authors"]}. <i>{paper["journal"]}</i>. {paper["year"]}.'
            story.append(Paragraph(citation_text, citation_style))
            
            link_parts = []
            if paper.get('pmid'):
                link_parts.append(
                    f'<link href="https://pubmed.ncbi.nlm.nih.gov/{paper["pmid"]}/" color="blue">'
                    f'PMID: {paper["pmid"]}</link>')
            if paper.get('doi'):
                link_parts.append(
                    f'<link href="https://doi.org/{paper["doi"]}" color="blue">'
                    f'DOI: {paper["doi"]}</link>')
            link_text = ' | '.join(link_parts) if link_parts else ''
            story.append(Paragraph(link_text, citation_style))
            story.append(Spacer(1, 10))
        
        # Build PDF
        doc.build(story)
        print(f"  PDF report saved: {filename}")
    
    def run(self):
        """Execute the complete search and report generation"""
        print("\n" + "="*80)
        print("SETBP1 LITERATURE SEARCH")
        print("="*80 + "\n")
        
        # Search PubMed
        pmids = self.search_pubmed()
        pubmed_papers = []
        if pmids:
            pubmed_papers = self.get_pubmed_metadata(pmids)
            self.all_papers.extend(pubmed_papers)

        # Search bioRxiv and medRxiv
        biorxiv_papers = self.search_biorxiv()
        medrxiv_papers = self.search_medrxiv()

        self.all_papers.extend(biorxiv_papers)
        self.all_papers.extend(medrxiv_papers)

        print(f"\nTotal papers found: {len(self.all_papers)}")
        print(f"  PubMed: {len(pubmed_papers)}")
        print(f"  bioRxiv: {len(biorxiv_papers)}")
        print(f"  medRxiv: {len(medrxiv_papers)}")

        if not self.all_papers:
            print("No papers found for this date range.")
            return

        # Categorize papers
        categories = self.categorize_papers(self.all_papers)

        # Generate output filenames with YYYYMMDD pattern
        end_date_formatted = self.end_date.replace('-', '')
        excel_filename = f"{self.output_dir}/{end_date_formatted}-SETBP1-Literature-Data.xlsx"
        pdf_filename = f"{self.output_dir}/{end_date_formatted}-SETBP1-Literature-Summary.pdf"

        # Create reports
        self.create_excel_report(self.all_papers, excel_filename)
        self.create_pdf_report(self.all_papers, categories, pdf_filename,
                              pubmed_count=len(pubmed_papers),
                              biorxiv_count=len(biorxiv_papers),
                              medrxiv_count=len(medrxiv_papers))
        
        print("\n" + "="*80)
        print("SEARCH COMPLETE")
        print("="*80)
        print(f"\nReports generated:")
        print(f"  Excel: {excel_filename}")
        print(f"  PDF:   {pdf_filename}")
        print(f"\nTotal papers: {len(self.all_papers)}")
        print(f"  Mechanism: {len(categories['mechanism'])}")
        print(f"  Therapeutics: {len(categories['therapeutics'])}")
        print(f"  Models: {len(categories['models'])}")
        print(f"  Datasets: {len(categories['dataset'])}")
        print(f"  Other: {len(categories['other'])}")


def main():
    """Main entry point"""
    parser = argparse.ArgumentParser(
        description='Search SETBP1 literature and generate reports',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Search last 7 days
  python setbp1_literature_search.py

  # Search specific date range
  python setbp1_literature_search.py --start 2026-02-01 --end 2026-02-08

  # Search with custom output directory
  python setbp1_literature_search.py --start 2026-02-01 --end 2026-02-08 --output /tmp
        """
    )
    
    parser.add_argument('--start', type=str,
                       help='Start date (YYYY-MM-DD). Default: 7 days ago')
    parser.add_argument('--end', type=str,
                       help='End date (YYYY-MM-DD). Default: today')
    parser.add_argument('--output', type=str, default='/mnt/user-data/outputs',
                       help='Output directory. Default: /mnt/user-data/outputs')
    
    args = parser.parse_args()
    
    # Set default dates
    if not args.end:
        args.end = datetime.now().strftime('%Y-%m-%d')
    
    if not args.start:
        end_date = datetime.strptime(args.end, '%Y-%m-%d')
        start_date = end_date - timedelta(days=7)
        args.start = start_date.strftime('%Y-%m-%d')
    
    # Validate dates
    try:
        datetime.strptime(args.start, '%Y-%m-%d')
        datetime.strptime(args.end, '%Y-%m-%d')
    except ValueError:
        print("Error: Dates must be in YYYY-MM-DD format")
        return 1
    
    # Run search
    searcher = SETBPLiteratureSearch(args.start, args.end, args.output)
    searcher.run()
    
    return 0


if __name__ == '__main__':
    exit(main())
